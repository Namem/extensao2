# -*- coding: utf-8 -*-
"""Gera o Backlog do Produto (Excel) para o Relatorio Final de Extensao II.
3 sprints (~1 mes cada), com priorizacao. Somente produto — sem TCC nem artigo."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "C:/Users/Rachid/Desktop/NR/Semestre 2026_1/extensao/ceres-diagnostico/docs/extensao/anexos/backlog_produto_ceres.xlsx"

VERDE = "1A3A1A"; CARD = "2D5A2D"; EPICO = "6E8B5A"
VOK = "E4EFE0"; VFUT = "FBEFD6"
PRIO = {"Crítica": "E6C7C2", "Alta": "F2E1BE", "Média": "D4E2F0", "Baixa": "E7E4DD",
        "Futura": "FBEFD6"}
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT = "Calibri"

wb = Workbook()

# ─────────────────────── ABA 1: Backlog ───────────────────────
ws = wb.active
ws.title = "Backlog Produto"
ws.merge_cells("A1:D1")
ws["A1"] = "Backlog do Produto — Ceres Diagnóstico (Extensão II — IFMT Cuiabá)"
ws["A1"].font = Font(name=FONT, size=14, bold=True, color=VERDE)
ws.row_dimensions[1].height = 24
ws.merge_cells("A2:D2")
ws["A2"] = ("3 sprints de ~1 mês · Prioridades: Crítica / Alta / Média / Baixa · "
            "Somente software, firmware e hardware (sem TCC ou artigo).")
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="777777")

HROW = 4
for c, h in enumerate(["Tarefa", "Camada", "Prioridade", "Status"], 1):
    cell = ws.cell(row=HROW, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=CARD)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    cell.border = BORDER

# (sprint_header, [(epico, [(tarefa, camada, prioridade, status), ...]), ...])
data = [
    ("Sprint 1 — Núcleo Inteligente (Backend + IA)   ·   ≈ 1 mês (4 semanas)   ·   ✅ 18/18", [
        ("Épico A — Motor de Diagnóstico e API (ex-Sprint 0)", [
            ("Mapeamento das 10 doenças do tomateiro (Embrapa)", "Backend", "Crítica", "Concluído"),
            ("Models Pergunta/Opcao/Diagnostico + migrations", "Backend", "Crítica", "Concluído"),
            ("Endpoints /iniciar/ e /responder/", "Backend", "Alta", "Concluído"),
            ("Autenticação JWT (SimpleJWT)", "Backend", "Alta", "Concluído"),
            ("Multi-tenant (Tenant + CustomUser)", "Backend", "Média", "Concluído"),
            ("PostgreSQL 18 (porta 5433)", "Infra", "Alta", "Concluído"),
            ("5 testes automatizados passando", "QA", "Alta", "Concluído"),
        ]),
        ("Épico B — Dataset, Modelo de IA e MQTT (ex-Sprint 1)", [
            ("PlantVillage (18.160 imgs) + augmentation → 88.949", "IA", "Crítica", "Concluído"),
            ("Exp A — Edge Impulse (FP32 92,5% / INT8 62%)", "IA", "Média", "Concluído"),
            ("Exp B — TF local (float 98,13% / INT8 95,76%, 639 KB)", "IA", "Crítica", "Concluído"),
            ("Exp C — Background augmentation sintética (20,24% campo)", "IA", "Baixa", "Concluído"),
            ("Exp D — Fine-tuning PlantDoc real (30,43% campo)", "IA", "Média", "Concluído"),
            ("Exp E — Focal Loss + aug agressiva (final, 638 KB)", "IA", "Crítica", "Concluído"),
            ("export_tflite.py — INT8 calibrado", "IA", "Alta", "Concluído"),
            ("Validação PlantDoc — gap lab-campo documentado", "IA", "Alta", "Concluído"),
            ("Backend MQTT (DiagnosticoEvento + mqtt_listener + /historico/)", "Backend", "Alta", "Concluído"),
            ("Mosquitto 2.1 instalado e testado", "Infra", "Média", "Concluído"),
            ("Matriz de confusão + acurácia por classe", "QA", "Média", "Concluído"),
        ]),
    ]),
    ("Sprint 2 — Borda Embarcada (Firmware + IoT)   ·   ≈ 1 mês (4 semanas)   ·   ✅ 8/8", [
        ("Épico A — Nó de sensores MQTT (ex-Sprint 1b)", [
            ("Projeto PlatformIO esp32_mqtt_sensor", "Firmware", "Alta", "Concluído"),
            ("WiFi + MQTT + reconexão automática", "Firmware", "Alta", "Concluído"),
            ("Publicação de JSON em ceres/sensor/", "Firmware", "Alta", "Concluído"),
            ("Pilha ESP32 → Mosquitto → Django → PostgreSQL", "Integração", "Alta", "Concluído"),
        ]),
        ("Épico B — TFLite Micro no ESP32-S3 (ex-Sprint 2)", [
            ("esp32s3_ceres — TFLite Micro (arena PSRAM 200 KB)", "Firmware", "Crítica", "Concluído"),
            ("gerar_arrays_c.py — modelo + 10 imgs como arrays C", "IA", "Alta", "Concluído"),
            ("Benchmark embarcado — 692 ms, 10/10 correto", "Firmware", "Crítica", "Concluído"),
            ("benchmark_esp32s3.md documentado", "QA", "Média", "Concluído"),
        ]),
    ]),
    ("Sprint 3 — Aplicativo e Integração (App + UX + Deploy)   ·   ≈ 1 mês (4 semanas)   ·   ✅ 29/29", [
        ("Épico A — App base + inferência (ex-Sprint 3)", [
            ("App Flutter (Diagnóstico + Histórico)", "App", "Crítica", "Concluído"),
            ("api_service — multipart POST + GET paginado", "App", "Alta", "Concluído"),
            ("View /inferir/ via subprocess TFLite", "Backend", "Crítica", "Concluído"),
            ("Validação end-to-end (galeria → TFLite → resultado)", "QA", "Alta", "Concluído"),
            ("Experimento Edge vs Cloud", "IA", "Média", "Concluído"),
            ("Drift (SQLite) — persistência local", "App", "Alta", "Concluído"),
        ]),
        ("Épico B — Design System e UI (ex-Sprint 3.5 + 3.6)", [
            ("CeresTheme (paleta cerrado, Material 3)", "App", "Alta", "Concluído"),
            ("6 telas redesenhadas (Splash/Login/Diagnóstico/IoT/Salvos/Enciclopédia)", "App", "Alta", "Concluído"),
            ("doencas_data.dart — 10 doenças centralizadas", "App", "Média", "Concluído"),
            ("flutter_svg + ícones thin-stroke", "App", "Baixa", "Concluído"),
            ("Tab bar + appbar fiéis ao mockup", "App", "Baixa", "Concluído"),
        ]),
        ("Épico C — Deploy, nuvem e novas telas (ex-Sprint 3.7)", [
            ("12 telas migradas + 4 novas (Alertas/Parceiro/Cadastro/Agrônomos)", "App", "Alta", "Concluído"),
            ("Registro de usuário /register/", "Backend", "Alta", "Concluído"),
            ("Deploy Railway — ceres.up.railway.app", "Infra", "Crítica", "Concluído"),
            ("TFLite on-device Android (~60 ms)", "App", "Crítica", "Concluído"),
            ("MQTT Cloud HiveMQ (TLS 8883 / WS 8884)", "IoT", "Alta", "Concluído"),
        ]),
        ("Épico D — UX: navegação, mapa e perfil (ex-Sprint 4A + 4B + 5)", [
            ("Back button no CeresAppBar", "App", "Baixa", "Concluído"),
            ("Persistência de sessão (shared_preferences) + auto-refresh JWT", "App", "Média", "Concluído"),
            ("Banner offline (connectivity_plus)", "App", "Média", "Concluído"),
            ("latitude/longitude em DiagnosticoEvento", "Backend", "Média", "Concluído"),
            ("MapaScreen (flutter_map + OSM, marcadores por urgência)", "App", "Alta", "Concluído"),
            ("Captura de GPS no diagnóstico", "App", "Média", "Concluído"),
            ("GET /api/auth/me/ (estatísticas do usuário)", "Backend", "Média", "Concluído"),
            ("PerfilScreen (stats, exportar CSV, logout)", "App", "Média", "Concluído"),
        ]),
        ("Épico E — Robustez e sincronização offline (ex-Sprint 5B)", [
            ("Railway PostgreSQL persistente entre deploys", "Infra", "Crítica", "Concluído"),
            ("Per-user diagnostics (FK usuario)", "Backend", "Alta", "Concluído"),
            ("Temperature scaling INT8 (T=0.25)", "Backend/App", "Alta", "Concluído"),
            ("Fila de sincronização offline → online (SyncService)", "App", "Alta", "Concluído"),
            ("Matriz de confusão INT8 (95,76%, 2.734 imgs)", "QA", "Média", "Concluído"),
        ]),
    ]),
    ("Próximos passos — Fase Futura (registrada)", [
        ("", [
            ("EfficientNet-B0 no Raspberry Pi 3B+", "IA", "Média", "Registrada"),
            ("Câmera OV5640 no ESP32-S3 (diagnóstico autônomo)", "Firmware", "Alta", "Registrada"),
        ]),
    ]),
]

row = HROW + 1
for sprint_header, epicos in data:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=sprint_header)
    c.font = Font(name=FONT, size=10.5, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=VERDE)
    c.alignment = Alignment(vertical="center", horizontal="left")
    c.border = BORDER
    ws.row_dimensions[row].height = 19
    row += 1
    for epico_nome, tarefas in epicos:
        if epico_nome:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            e = ws.cell(row=row, column=1, value="   " + epico_nome)
            e.font = Font(name=FONT, size=9.5, bold=True, italic=True, color="FFFFFF")
            e.fill = PatternFill("solid", fgColor=EPICO)
            e.alignment = Alignment(vertical="center", horizontal="left")
            e.border = BORDER
            row += 1
        for tarefa, camada, prioridade, status in tarefas:
            fill_status = VOK if status == "Concluído" else VFUT
            for col, v in enumerate([tarefa, camada, prioridade, status], 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = Font(name=FONT, size=9, color="261E19", bold=(col in (3, 4)))
                cell.border = BORDER
                if col == 1:
                    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                if col == 3:
                    cell.fill = PatternFill("solid", fgColor=PRIO.get(prioridade, "FFFFFF"))
                if col == 4:
                    cell.fill = PatternFill("solid", fgColor=fill_status)
            row += 1

for i, w in enumerate([64, 13, 12, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ─────────────────────── ABA 2: Resumo ───────────────────────
ws2 = wb.create_sheet("Resumo")
ws2.merge_cells("A1:F1")
ws2["A1"] = "Resumo do Backlog do Produto — Ceres Diagnóstico"
ws2["A1"].font = Font(name=FONT, size=14, bold=True, color=VERDE)
ws2.row_dimensions[1].height = 24

for c, h in enumerate(["Sprint", "Tema", "Duração", "Camada", "Status", "Tarefas"], 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=CARD)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    cell.border = BORDER

resumo = [
    ("Sprint 1", "Núcleo Inteligente (Backend + IA)", "≈ 1 mês", "Backend / IA", "Concluída", "18/18"),
    ("Sprint 2", "Borda Embarcada (Firmware + IoT)", "≈ 1 mês", "Firmware / IoT", "Concluída", "8/8"),
    ("Sprint 3", "Aplicativo e Integração (App + UX + Deploy)", "≈ 1 mês", "App / Infra", "Concluída", "29/29"),
    ("Fase Futura", "RPi3B+ + câmera OV5640", "—", "IA / Firmware", "Registrada", "—"),
]
r = 4
for sprint, tema, dur, camada, status, tarefas in resumo:
    fill = VOK if status == "Concluída" else VFUT
    for col, v in enumerate([sprint, tema, dur, camada, status, tarefas], 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.font = Font(name=FONT, size=9, color="261E19", bold=(col in (1, 5)))
        cell.alignment = Alignment(vertical="center", horizontal="center" if col in (3, 5, 6) else "left")
        cell.border = BORDER
        if col == 5:
            cell.fill = PatternFill("solid", fgColor=fill)
    r += 1

ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws2.cell(row=r, column=1, value="TOTAL (produto)").font = Font(name=FONT, size=10, bold=True, color=VERDE)
ws2.cell(row=r, column=1).alignment = Alignment(horizontal="right")
tot = ws2.cell(row=r, column=6, value="55/55")
tot.font = Font(name=FONT, size=10, bold=True, color=VERDE)
tot.alignment = Alignment(horizontal="center")

for i, w in enumerate([13, 42, 20, 16, 13, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print("OK:", OUT)
